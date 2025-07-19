#!/usr/bin/env python3
import json
from datetime import datetime
import os

# 샘플 데이터 생성
sample_data = [
    {
        '로스터리': '프릳츠커피',
        '커피명': '에티오피아 예가체프',
        '원산지': '에티오피아 예가체프',
        '가공방식': 'Washed',
        '테이스팅노트': '플로럴, 레몬, 자스민',
        '가격': 18000,
        '중량': '200g',
        '로스팅레벨': 'Light',
        '품종': 'Heirloom',
        '고도': '1900-2100m',
        '수확시기': '2023',
        'URL': 'https://fritz.co.kr/product/ethiopia',
        '크롤링시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    },
    {
        '로스터리': '센터커피',
        '커피명': '콜롬비아 핑크 버번',
        '원산지': '콜롬비아 우일라',
        '가공방식': 'Honey',
        '테이스팅노트': '체리, 카카오, 브라운슈가',
        '가격': 22000,
        '중량': '200g',
        '로스팅레벨': 'Medium',
        '품종': 'Pink Bourbon',
        '고도': '1700m',
        '수확시기': '2023',
        'URL': 'https://centercoffee.kr/product/colombia',
        '크롤링시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    },
    {
        '로스터리': '테라로사',
        '커피명': '과테말라 안티구아',
        '원산지': '과테말라 안티구아',
        '가공방식': 'Washed',
        '테이스팅노트': '초콜릿, 오렌지, 캐러멜',
        '가격': 19000,
        '중량': '200g',
        '로스팅레벨': 'Medium',
        '품종': 'Bourbon, Caturra',
        '고도': '1500-1700m',
        '수확시기': '2023',
        'URL': 'https://terarosa.com/product/guatemala',
        '크롤링시간': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    }
]

# CSV 파일로 저장
import csv

os.makedirs('data/processed', exist_ok=True)

csv_file = f'data/processed/all_coffee_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'

with open(csv_file, 'w', newline='', encoding='utf-8-sig') as f:
    writer = csv.DictWriter(f, fieldnames=sample_data[0].keys())
    writer.writeheader()
    writer.writerows(sample_data)

print(f"샘플 데이터가 {csv_file}에 저장되었습니다.")
print(f"총 {len(sample_data)}개의 커피 정보가 저장되었습니다.")

# Excel 파일로도 저장
try:
    import openpyxl
    from openpyxl import Workbook
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Coffee Data"
    
    # 헤더 작성
    headers = list(sample_data[0].keys())
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)
    
    # 데이터 작성
    for row, data in enumerate(sample_data, 2):
        for col, key in enumerate(headers, 1):
            ws.cell(row=row, column=col, value=data[key])
    
    # 열 너비 조정
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width
    
    excel_file = f'data/processed/all_coffee_data_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    wb.save(excel_file)
    print(f"Excel 파일도 {excel_file}에 저장되었습니다.")
except:
    print("Excel 파일 저장은 openpyxl 패키지가 필요합니다.")